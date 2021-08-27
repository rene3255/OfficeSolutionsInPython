from os import readlink
from tkinter import Message
from tkinter.constants import X
from openpyxl.worksheet.pagebreak import Break
from openpyxl import workbook
from openpyxl import load_workbook
import termcolor
import os

wb = load_workbook('nd_2020.xlsx')
ws = wb['ND_2020']
ws = wb.active

def queryVariante(strCellVal, filas, rango):
    cell_range = rango
    filas_fol = filas
    iCount = -1
    result = ''
    for i in filas_fol:
        iCount +=1
        if str(i[0].value).find(strCellVal.upper()) != -1:
            result +="".join(f'Remesa {cell_range[iCount][12].value} | Folio de Captura {cell_range[iCount][4].value} | ND no. {cell_range[iCount][5].value} | Estatus {cell_range[iCount][13].value}\n{cell_range[iCount][9].value}')+'\n'
    
    return result

isNumeric = lambda y:  False if list(filter(lambda x: x not in '0123456789' ,y)) else True

def queryExcel(strCellVal,filas, rango):
    cell_range = rango
    filas_fol = filas
    result = ''
    iCount = -1
   
    if isNumeric(strCellVal) and len(str(strCellVal)) == 13: # Busqueda por folio de captura
        for i in filas_fol:
            iCount +=1
            

            if str(i[0].value) == str(strCellVal):
               
                result = f'La remesa es {cell_range[iCount][12].value} con estatus {cell_range[iCount][13].value} pertenece a {cell_range[iCount][9].value}'
                break
    elif isNumeric(strCellVal) and len(str(strCellVal)) == 10: # Por Numero de ND
        for i in filas_fol:
            
            iCount +=1
            if str(i[0].value) == str(strCellVal):
                result = f'La remesa es {cell_range[iCount][12].value} con estatus {cell_range[iCount][13].value} pertenece a {cell_range[iCount][9].value}'
                break
    else: # Por Nombre 
        for i in filas_fol:
            iCount +=1
            if i[0].value == strCellVal.upper():
                result = f'La remesa es {cell_range[iCount][12].value} con estatus {cell_range[iCount][13].value} con numero de formato ND {cell_range[iCount][5].value}'
                break

            
    string = "No-se-encontraron-coincidencias"
    string = termcolor.colored(string, 'red')
    return result if result  else string

#print(queryExcel('1010620001245',ws['E2':'E15'],ws['A2':'O16']))
miMenu = 0
msgError = termcolor.colored('No es una opciòn valida', 'green')
msgError2 = termcolor.colored('debe der entrada nùmerica', 'blue')

os.system('clear')
while miMenu != 5:
    print('[1.- Nombre]')
    print('[2.- de Folio de Captura]')
    print('[3.- Folio ND]')
    print('[4.- Variante alfabetica:')
    print('[5.- Salir]')
    try:
        miMenu = int(input('Seleccion: '))
       
        if miMenu == 1:
            try:
                miconsulta = str(input('Nombre:'))
            except ValueError:
                print(msgError2)

            print(queryExcel(miconsulta,ws['J2':'J6167'],ws['A2':'N6167']))
    
        elif miMenu == 2:
            try:
            
                miconsulta = input('No. de Folio:')

                print(queryExcel(miconsulta,ws['E2':'E6167'],ws['A2':'N6167']))
            except ValueError:
                print(msgError2)

        elif miMenu == 3:
            miconsulta = input('No. de ND: ')
            print(queryExcel(miconsulta,ws['F2':'F6167'],ws['A2':'N6167']))
        elif miMenu == 4:
            miconsulta = input('Variante alfabètica: ')
            print(queryVariante(miconsulta,ws['J2':'J6167'],ws['A2':'N6167']))
        elif miMenu == 6:
            os.system('Clear')


             
    except ValueError:
        os.system('clear')
        print(msgError)